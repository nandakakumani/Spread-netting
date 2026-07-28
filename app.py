
import io
import re
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional

import pandas as pd
import streamlit as st

# ============================================================
# Default COB date
# ============================================================

default_cob_date = (
        pd.Timestamp.today().normalize() - pd.tseries.offsets.BDay(1)
).date()


# ============================================================
# 1) Netting core logic
# ============================================================

def sgn(x: float) -> int:
    return 0 if abs(x) < 1e-12 else (1 if x > 0 else -1)


@dataclass
class NettingConfig:
    hubs_eu: List[str]
    hubs_asia: List[str]


def net_pair(pos: Dict[str, float], spreads: Dict[str, float], A: str, B: str) -> float:
    a = float(pos.get(A, 0.0))
    b = float(pos.get(B, 0.0))

    if sgn(a) == 0 or sgn(b) == 0:
        return 0.0

    if sgn(a) == sgn(b):
        return 0.0

    x = min(abs(a), abs(b))
    spread_val = sgn(a) * x

    key = f"{A}/ {B}"
    spreads[key] = spreads.get(key, 0.0) + spread_val

    pos[A] = a - spread_val
    pos[B] = b + spread_val

    return x


def run_netting_staged(
        outrights: Dict[str, float],
        cfg: NettingConfig,
) -> Tuple[
    Dict[str, float],
    Dict[str, float],
    Dict[str, float],
    Dict[str, float],
    Dict[str, float],
    Dict[str, float],
]:
    pos = {k: float(v) for k, v in outrights.items()}

    spreads_eu: Dict[str, float] = {}
    spreads_asia: Dict[str, float] = {}
    spreads_cross: Dict[str, float] = {}

    eu = cfg.hubs_eu
    asia = cfg.hubs_asia

    # EU internal
    for i in range(len(eu)):
        A = eu[i]
        for j in range(i + 1, len(eu)):
            B = eu[j]
            net_pair(pos, spreads_eu, A, B)

    pos_after_eu = dict(pos)

    # Asia triangle
    if "INDIA" in asia and "FE" in asia:
        net_pair(pos, spreads_asia, "INDIA", "FE")
    if "INDIA" in asia and "JKM" in asia:
        net_pair(pos, spreads_asia, "INDIA", "JKM")
    if "FE" in asia and "JKM" in asia:
        net_pair(pos, spreads_asia, "FE", "JKM")

    pos_after_asia = dict(pos)

    # EU vs Asia
    for A in eu:
        for B in asia:
            net_pair(pos, spreads_cross, A, B)

    pos_final = dict(pos)

    return pos_after_eu, spreads_eu, pos_after_asia, spreads_asia, pos_final, spreads_cross


def spread_groups(cfg: NettingConfig) -> Tuple[List[str], List[str], List[str]]:
    eu = cfg.hubs_eu
    asia = cfg.hubs_asia

    intra_eu = [
        f"{eu[i]}/ {eu[j]}"
        for i in range(len(eu))
        for j in range(i + 1, len(eu))
    ]

    asia_tri_all = ["INDIA/ FE", "INDIA/ JKM", "FE/ JKM"]
    asia_tri = []

    for c in asia_tri_all:
        a, b = c.replace("/ ", "/").split("/")
        if a in asia and b in asia:
            asia_tri.append(c)

    eu_asia = [f"{a}/ {b}" for a in eu for b in asia]

    return intra_eu, asia_tri, eu_asia


# ============================================================
# 2) Input constants and helpers
# ============================================================

IFRS_SHEETS = {
    "IFRS 26": "IFRS 26 discounted",
    "IFRS 27": "IFRS 27 discounted",
    "IFRS 28": "IFRS 28 discounted",
}

SOURCE_YEAR_CELL = "D1"
PRODUCT_HEADER_ROW = 8
DATA_START_ROW = 9

EXCLUDE_FROM_OUTPUT_PRODUCTS = {
    "EUA",
}

FREIGHT_PRODUCTS = [
    "2-Stroke Freight",
    "Steam Freight",
    "TFDE Freight",
]

PIVOT_PRODUCT_RENAME = {
    "LNG H DES India": "INDIA",
    "LNG H DES JKM": "FE",
    "LNG H DES NWE": "NWE",
    "LNG H DES Med": "ATL",
    "PEG - DNK": "PEG",
    "ZTP - DNK": "ZTP",
}

SCALE_DIVIDE_BY_MILLION = {
    "HH",
    "Brent Bullet",
    "Brent Futures",
    "Brent Swaps",
    "Dated Brent",
    "JCC",
    "Dubai",
}

RAW_OTHER_PRODUCT_ORDER = [
    "HH",
    "Brent Bullet",
    "Brent Futures",
    "Brent Swaps",
    "Dated Brent",
    "JCC",
    "Dubai",
]


def _clean_header(x: str) -> str:
    return str(x).replace("\u00A0", " ").strip()


def _canonical_header(x: str) -> str:
    x = _clean_header(x)
    return re.sub(r"\.\d+$", "", x)


def _norm_name(x) -> str:
    if x is None:
        return ""
    return str(x).replace("\u00A0", " ").strip()


def canonical_product_name(x: str) -> str:
    x = str(x).replace("\u00A0", " ").strip()
    x = x.replace("–", "-").replace("—", "-")
    x = re.sub(r"\s*-\s*", "-", x)
    x = re.sub(r"\s+", " ", x)
    return x.lower()


FREIGHT_PRODUCTS_CANONICAL = {
    canonical_product_name(x) for x in FREIGHT_PRODUCTS
}


def is_freight_product(x: str) -> bool:
    return canonical_product_name(x) in FREIGHT_PRODUCTS_CANONICAL


def normalise_freight_display_name(x: str) -> str:
    cx = canonical_product_name(x)

    for p in FREIGHT_PRODUCTS:
        if canonical_product_name(p) == cx:
            return p

    return x


def normalize_pivot_product_name(name) -> str:
    x = _norm_name(name)
    x = x.replace("–", "-").replace("—", "-")
    x = re.sub(r"\s*-\s*", " - ", x)
    x = re.sub(r"\s+", " ", x).strip()
    return x


def _is_non_netted_product(x: str) -> bool:
    return is_freight_product(x)


def dedupe_preserve_order(items: List[str]) -> List[str]:
    seen = set()
    out = []

    for x in items:
        if x not in seen:
            out.append(x)
            seen.add(x)

    return out


def normalize_and_merge_duplicate_columns(df: pd.DataFrame, date_col_name: str) -> pd.DataFrame:
    if df.empty:
        return df

    original_cols = list(df.columns)
    cleaned_cols = [_clean_header(c) for c in original_cols]
    canon_cols = [_canonical_header(c) for c in original_cols]

    df = df.copy()
    df.columns = cleaned_cols

    cleaned_to_canon = {
        cleaned_cols[i]: canon_cols[i]
        for i in range(len(cleaned_cols))
    }

    canon_list = [cleaned_to_canon[c] for c in df.columns]

    if len(set(canon_list)) == len(canon_list):
        df.columns = canon_list
        return df

    out = pd.DataFrame(index=df.index)
    date_clean = _clean_header(date_col_name)
    date_canon = _canonical_header(date_col_name)

    groups: Dict[str, List[str]] = {}

    for c in df.columns:
        groups.setdefault(cleaned_to_canon[c], []).append(c)

    for canon_name, cols in groups.items():
        if canon_name == date_clean or canon_name == date_canon:
            out[canon_name] = df[cols].bfill(axis=1).iloc[:, 0]
        else:
            block = df[cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
            out[canon_name] = block.sum(axis=1)

    return out


def extract_year_from_cell_value(x) -> Optional[int]:
    if x is None or pd.isna(x):
        return None

    dt = pd.to_datetime(x, errors="coerce", dayfirst=True)

    if pd.notna(dt):
        return int(dt.year)

    num = pd.to_numeric(pd.Series([x]), errors="coerce").iloc[0]

    if pd.notna(num) and 20000 < num < 80000:
        dt = pd.to_datetime(
            num,
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )

        if pd.notna(dt):
            return int(dt.year)

    m = re.search(r"(19\d{2}|20\d{2})", str(x))
    return int(m.group(1)) if m else None


def month_text_to_number(x) -> Optional[int]:
    if x is None or pd.isna(x):
        return None

    text = str(x).strip()
    dt = pd.to_datetime(f"1 {text} 2000", errors="coerce")

    if pd.notna(dt):
        return int(dt.month)

    return None


def find_sheet_case_insensitive(workbook, target_name: str) -> str:
    sheet_lookup = {s.lower().strip(): s for s in workbook.sheetnames}
    key = target_name.lower().strip()

    if key not in sheet_lookup:
        raise ValueError(
            f"Sheet '{target_name}' not found. Available sheets: {workbook.sheetnames}"
        )

    return sheet_lookup[key]


def read_ifrs_sheet_input(
        file,
        sheet_name: str,
) -> Tuple[pd.DataFrame, Optional[int], List[str], Dict[str, str], str]:
    from openpyxl import load_workbook

    file.seek(0)
    wb = load_workbook(file, data_only=True, read_only=True)

    actual_sheet_name = find_sheet_case_insensitive(wb, sheet_name)
    ws = wb[actual_sheet_name]

    source_year = extract_year_from_cell_value(ws[SOURCE_YEAR_CELL].value)

    product_cols = []
    original_products = []
    rename_used = {}

    for col in range(5, ws.max_column + 1):
        original_name = normalize_pivot_product_name(ws.cell(PRODUCT_HEADER_ROW, col).value)

        if not original_name:
            continue

        if original_name.lower() == "grand total":
            continue

        if original_name in EXCLUDE_FROM_OUTPUT_PRODUCTS:
            continue

        original_products.append(original_name)

        if is_freight_product(original_name):
            output_name = normalise_freight_display_name(original_name)
        else:
            output_name = PIVOT_PRODUCT_RENAME.get(original_name, original_name)

        rename_used[original_name] = output_name
        product_cols.append((col, original_name, output_name))

    rows = []

    for row in range(DATA_START_ROW, ws.max_row + 1):
        year_val = ws.cell(row, 3).value
        month_val = ws.cell(row, 4).value

        if year_val is None and month_val is None:
            continue

        if str(year_val).strip().lower() == "grand total":
            continue

        year_num = pd.to_numeric(pd.Series([year_val]), errors="coerce").iloc[0]
        month_num = month_text_to_number(month_val)

        if pd.isna(year_num) or month_num is None:
            continue

        rec = {
            "Date": pd.Timestamp(
                year=int(year_num),
                month=int(month_num),
                day=1,
            )
        }

        for col, original_name, output_name in product_cols:
            raw_val = ws.cell(row, col).value
            num = pd.to_numeric(pd.Series([raw_val]), errors="coerce").iloc[0]
            value = float(num) if pd.notna(num) else 0.0

            if original_name in SCALE_DIVIDE_BY_MILLION or output_name in SCALE_DIVIDE_BY_MILLION:
                value = value / 1_000_000

            rec[output_name] = rec.get(output_name, 0.0) + value

        rows.append(rec)

    df = pd.DataFrame(rows)

    if df.empty:
        return df, source_year, original_products, rename_used, actual_sheet_name

    df = normalize_and_merge_duplicate_columns(df, date_col_name="Date")

    return df, source_year, original_products, rename_used, actual_sheet_name


def ensure_columns(df: pd.DataFrame, required_cols: List[str]) -> pd.DataFrame:
    out = df.copy()

    for c in required_cols:
        if c not in out.columns:
            out[c] = 0.0

    return out


def parse_mixed_excel_dates(s: pd.Series, dayfirst: bool = False) -> pd.Series:
    x = s.copy()
    as_num = pd.to_numeric(x, errors="coerce")

    mask_excel = as_num.notna() & (as_num > 20000) & (as_num < 80000)

    out = pd.Series(pd.NaT, index=x.index, dtype="datetime64[ns]")

    if mask_excel.any():
        out.loc[mask_excel] = pd.to_datetime(
            as_num.loc[mask_excel],
            unit="D",
            origin="1899-12-30",
            errors="coerce",
        )

    mask_other = ~mask_excel

    if mask_other.any():
        out.loc[mask_other] = pd.to_datetime(
            x.loc[mask_other],
            errors="coerce",
            dayfirst=dayfirst,
        )

    return out


def drop_all_zero_periods(df: pd.DataFrame, eps: float = 1e-12) -> pd.DataFrame:
    if df.empty:
        return df

    numeric = df.select_dtypes(include="number")

    if numeric.empty:
        return df

    keep = numeric.abs().sum(axis=1) >= eps
    return df.loc[keep]


def monthly_positions_from_input(
        df: pd.DataFrame,
        date_col: str,
        hubs: List[str],
        dayfirst: bool = False,
) -> pd.DataFrame:
    out = df.copy()

    out[date_col] = parse_mixed_excel_dates(out[date_col], dayfirst=dayfirst)
    out = out.dropna(subset=[date_col])

    for h in hubs:
        out[h] = pd.to_numeric(out[h], errors="coerce").fillna(0.0)

    out = out.loc[out[hubs].abs().sum(axis=1) >= 1e-12]
    out["Month"] = out[date_col].dt.to_period("M")

    monthly = out.groupby("Month")[hubs].sum()
    monthly.index.name = "Period"
    monthly = drop_all_zero_periods(monthly)

    return monthly


def detect_other_products(df: pd.DataFrame, date_col: str, nettable: List[str]) -> List[str]:
    candidates = [
        c for c in df.columns
        if c != date_col
           and c not in nettable
           and c not in EXCLUDE_FROM_OUTPUT_PRODUCTS
    ]

    other = []

    for c in candidates:
        s = pd.to_numeric(df[c], errors="coerce").fillna(0.0)

        if s.abs().sum() >= 1e-12:
            other.append(c)

    return other


# ============================================================
# 3) Netting outputs
# ============================================================

def netting_outputs_for_periods_staged(
        positions: pd.DataFrame,
        cfg: NettingConfig,
) -> Dict[str, pd.DataFrame]:
    net_hubs = cfg.hubs_eu + cfg.hubs_asia
    passthrough = [c for c in positions.columns if c not in net_hubs]

    intra_cols, tri_cols, cross_cols = spread_groups(cfg)

    outrights = positions.copy()

    for h in net_hubs:
        if h not in outrights.columns:
            outrights[h] = 0.0

    all_cols = net_hubs + passthrough
    outrights = outrights.reindex(columns=all_cols, fill_value=0.0)

    eu_spread_rows = []
    asia_spread_rows = []
    cross_spread_rows = []

    pos_after_eu_rows = []
    pos_after_asia_rows = []
    pos_final_rows = []

    for period, row in outrights.iterrows():
        net_dict = {h: float(row.get(h, 0.0)) for h in net_hubs}
        pass_dict = {c: float(row.get(c, 0.0)) for c in passthrough}

        pos_eu, spr_eu, pos_asia, spr_asia, pos_final, spr_cross = run_netting_staged(
            net_dict,
            cfg,
        )

        pos_eu.update(pass_dict)
        pos_asia.update(pass_dict)
        pos_final.update(pass_dict)

        eu_row = {c: 0.0 for c in intra_cols}
        asia_row = {c: 0.0 for c in tri_cols}
        cross_row = {c: 0.0 for c in cross_cols}

        for k, v in spr_eu.items():
            if k in eu_row:
                eu_row[k] = float(v)

        for k, v in spr_asia.items():
            if k in asia_row:
                asia_row[k] = float(v)

        for k, v in spr_cross.items():
            if k in cross_row:
                cross_row[k] = float(v)

        eu_spread_rows.append(pd.Series(eu_row, name=period))
        asia_spread_rows.append(pd.Series(asia_row, name=period))
        cross_spread_rows.append(pd.Series(cross_row, name=period))

        pos_after_eu_rows.append(pd.Series(pos_eu, name=period))
        pos_after_asia_rows.append(pd.Series(pos_asia, name=period))
        pos_final_rows.append(pd.Series(pos_final, name=period))

    eu_spreads_df = (
        pd.DataFrame(eu_spread_rows)
        .fillna(0.0)
        .reindex(columns=intra_cols, fill_value=0.0)
    )

    asia_spreads_df = (
        pd.DataFrame(asia_spread_rows)
        .fillna(0.0)
        .reindex(columns=tri_cols, fill_value=0.0)
    )

    cross_spreads_df = (
        pd.DataFrame(cross_spread_rows)
        .fillna(0.0)
        .reindex(columns=cross_cols, fill_value=0.0)
    )

    pos_after_eu_df = (
        pd.DataFrame(pos_after_eu_rows)
        .fillna(0.0)
        .reindex(columns=all_cols, fill_value=0.0)
    )

    pos_after_asia_df = (
        pd.DataFrame(pos_after_asia_rows)
        .fillna(0.0)
        .reindex(columns=all_cols, fill_value=0.0)
    )

    residual_df = (
        pd.DataFrame(pos_final_rows)
        .fillna(0.0)
        .reindex(columns=all_cols, fill_value=0.0)
    )

    return {
        "outrights": drop_all_zero_periods(outrights),
        "eu_spreads": drop_all_zero_periods(eu_spreads_df),
        "pos_after_eu": drop_all_zero_periods(pos_after_eu_df),
        "asia_spreads": drop_all_zero_periods(asia_spreads_df),
        "pos_after_asia": drop_all_zero_periods(pos_after_asia_df),
        "cross_spreads": drop_all_zero_periods(cross_spreads_df),
        "residual": drop_all_zero_periods(residual_df),
    }


# ============================================================
# 4) Units and Streamlit formatting
# ============================================================

UNIT_FACTORS = {
    "MMBtu": 1.0,
    "TBtu": 1e-6,
}


def blank_zeros(df: pd.DataFrame, eps: float = 1e-12) -> pd.DataFrame:
    return df.mask(df.abs() < eps, pd.NA)


def _fmt_value(x, unit: str) -> str:
    if pd.isna(x):
        return ""

    if abs(float(x)) < 1e-12:
        return ""

    if unit == "TBtu":
        return f"({abs(x):,.2f})" if x < 0 else f"{x:,.2f}"

    return f"({abs(x):,.0f})" if x < 0 else f"{x:,.0f}"


def style_table_mixed(
        df: pd.DataFrame,
        unit_for_netting: str,
        passthrough_cols: Optional[List[str]] = None,
):
    passthrough_cols = passthrough_cols or []

    def fmt_for_col(col):
        if col in passthrough_cols:
            return lambda x: _fmt_value(x, "TBtu")
        return lambda x: _fmt_value(x, unit_for_netting)

    formatters = {c: fmt_for_col(c) for c in df.columns}

    def color_neg(v):
        try:
            if pd.notna(v) and float(v) < 0:
                return "color: #c00000; font-weight: 600;"
        except Exception:
            pass
        return ""

    return df.style.format(formatters, na_rep="").map(color_neg)


def nonzero_only_row(s: pd.Series, eps: float = 1e-12) -> pd.Series:
    return s[s.abs() >= eps].sort_values(key=lambda x: x.abs(), ascending=False)


def convert_stage_selective(
        stage: Dict[str, pd.DataFrame],
        cfg: NettingConfig,
        unit: str,
) -> Dict[str, pd.DataFrame]:
    net_cols = cfg.hubs_eu + cfg.hubs_asia
    out: Dict[str, pd.DataFrame] = {}

    for k, df in stage.items():
        d = df.copy()

        if k in {"outrights", "pos_after_eu", "pos_after_asia", "residual"}:
            hubs_present = [c for c in net_cols if c in d.columns]
            passthrough_present = [c for c in d.columns if c not in net_cols]

            if unit == "TBtu" and hubs_present:
                d.loc[:, hubs_present] = d[hubs_present] * UNIT_FACTORS["TBtu"]

            d = d.reindex(columns=hubs_present + passthrough_present)

        else:
            if unit == "TBtu":
                d = d * UNIT_FACTORS["TBtu"]

        out[k] = d

    return out


# ============================================================
# 5) Excel export
# ============================================================

def export_to_excel_bytes_staged(
        monthly: Dict[str, pd.DataFrame],
        unit: str,
        cfg: NettingConfig,
        ifrs_label: str,
        source_sheet_name: str,
        cob_date_str: str,
        source_year: Optional[int] = None,
        source_products: Optional[List[str]] = None,
        rename_used: Optional[Dict[str, str]] = None,
) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # -------------------------
    # Number formats
    # -------------------------

    def excel_number_format_local(unit_name: str) -> str:
        if unit_name == "TBtu":
            return '#,##0.00;[Red](#,##0.00);"-"'
        return '#,##0;[Red](#,##0);"-"'

    num_fmt_net = excel_number_format_local(unit)
    num_fmt_passthrough = '#,##0.00;[Red](#,##0.00);"-"'

    # -------------------------
    # Styles
    # -------------------------

    yellow_fill = PatternFill("solid", fgColor="FFF9C4")
    orange_fill = PatternFill("solid", fgColor="FCE4D6")
    raw_fill = PatternFill("solid", fgColor="FFF2CC")
    input_ratio_fill = PatternFill("solid", fgColor="C9C2E0")
    net_eq_fill = PatternFill("solid", fgColor="DDEBF7")
    header_fill = PatternFill("solid", fgColor="F2F2F2")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    title_fill = PatternFill("solid", fgColor="FFFFFF")
    stale_period_fill = PatternFill("solid", fgColor="FF0000")

    title_font = Font(bold=True, size=12, color="000000")
    section_font = Font(bold=True, size=9, color="000000")
    header_font = Font(bold=True, size=8, color="000000")
    normal_font = Font(size=8, color="000000")
    total_font = Font(bold=True, size=8, color="000000")
    period_font = Font(size=8, color="000000")
    stale_period_font = Font(bold=True, size=8, color="FFFFFF")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    thin = Side(style="thin", color="D9D9D9")
    medium = Side(style="medium", color="000000")
    none_side = Side(style=None)

    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # -------------------------
    # Generic helpers
    # -------------------------

    def set_col_widths(ws, max_col: int, default: float = 8.5):
        for col_num in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = default
        ws.column_dimensions["A"].width = 12

    def outline_border(ws, r1, c1, r2, c2):
        if r2 < r1 or c2 < c1:
            return

        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                ws.cell(row=r, column=c).border = thin_border

        for c in range(c1, c2 + 1):
            ws.cell(row=r1, column=c).border = Border(
                left=ws.cell(row=r1, column=c).border.left,
                right=ws.cell(row=r1, column=c).border.right,
                top=medium,
                bottom=ws.cell(row=r1, column=c).border.bottom,
            )
            ws.cell(row=r2, column=c).border = Border(
                left=ws.cell(row=r2, column=c).border.left,
                right=ws.cell(row=r2, column=c).border.right,
                top=ws.cell(row=r2, column=c).border.top,
                bottom=medium,
            )

        for r in range(r1, r2 + 1):
            ws.cell(row=r, column=c1).border = Border(
                left=medium,
                right=ws.cell(row=r, column=c1).border.right,
                top=ws.cell(row=r, column=c1).border.top,
                bottom=ws.cell(row=r, column=c1).border.bottom,
            )
            ws.cell(row=r, column=c2).border = Border(
                left=ws.cell(row=r, column=c2).border.left,
                right=medium,
                top=ws.cell(row=r, column=c2).border.top,
                bottom=ws.cell(row=r, column=c2).border.bottom,
            )

    def normalize_for_export(stage: Dict[str, pd.DataFrame]) -> Dict[str, pd.DataFrame]:
        base = stage["outrights"].copy()
        base.index = base.index.astype(str)

        master = pd.Index(list(base.index), name="Period")
        master_plus_total = pd.Index(list(master) + ["TOTAL"], name="Period")

        out = {}

        for k, df in stage.items():
            d = df.copy()
            d.index = d.index.astype(str)

            d_no_total = d.drop(index="TOTAL", errors="ignore")
            d_no_total = d_no_total.reindex(master, fill_value=0.0)
            d_no_total.loc["TOTAL"] = pd.NA

            out[k] = d_no_total

        out["_master_index"] = pd.DataFrame(index=master_plus_total)

        return out

    def inject_total_formulas(ws, df, start_row, start_col):
        if df.empty:
            return

        periods = list(df.index.astype(str))

        if "TOTAL" not in periods:
            return

        total_idx = periods.index("TOTAL")
        first_data_row = start_row + 2
        total_row = start_row + 2 + total_idx
        last_data_row = total_row - 1

        if last_data_row < first_data_row:
            return

        for j in range(len(df.columns)):
            col_letter = get_column_letter(start_col + j)
            ws.cell(row=total_row, column=start_col + j).value = (
                f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})"
            )

    def style_total_row(ws, df, start_row, start_col):
        if df.empty:
            return

        periods = list(df.index.astype(str))

        if "TOTAL" not in periods:
            return

        total_idx = periods.index("TOTAL")
        total_row = start_row + 2 + total_idx

        for j in range(len(df.columns)):
            cell = ws.cell(row=total_row, column=start_col + j)
            cell.font = total_font
            cell.fill = white_fill
            cell.border = Border(
                top=medium,
                bottom=none_side,
                left=none_side,
                right=none_side,
            )

    def write_period_column(ws, start_row: int, start_col: int, periods: List[str]) -> int:
        cell = ws.cell(start_row, start_col, "Period")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

        expected_month = pd.Period(pd.to_datetime(cob_date_str), freq="M")

        for i, p in enumerate(periods, start=1):
            row_cell = ws.cell(start_row + i, start_col, p)
            row_cell.alignment = left
            row_cell.fill = white_fill
            row_cell.font = period_font
            row_cell.border = thin_border

            if str(p).upper() == "TOTAL":
                row_cell.font = total_font
                continue

            try:
                row_period = pd.Period(str(p), freq="M")

                if row_period < expected_month:
                    row_cell.fill = stale_period_fill
                    row_cell.font = stale_period_font

            except Exception:
                pass

        return start_row + len(periods)

    def write_box(
            ws,
            title: str,
            df: pd.DataFrame,
            start_row: int,
            start_col: int,
            default_num_fmt: Optional[str] = None,
            data_fill: Optional[PatternFill] = None,
            formula_map: Optional[Dict[str, callable]] = None,
            outline: bool = True,
    ) -> Tuple[int, int]:

        default_num_fmt = default_num_fmt or num_fmt_net
        data_fill = data_fill or yellow_fill
        formula_map = formula_map or {}

        cols = list(df.columns)

        if len(cols) == 0:
            return start_row, start_col

        ncols = len(cols)
        periods = list(df.index.astype(str))
        nrows = len(periods)

        title_cell = ws.cell(start_row, start_col, title)
        title_cell.font = section_font
        title_cell.fill = title_fill
        title_cell.alignment = center
        title_cell.border = thin_border

        if ncols > 1:
            ws.merge_cells(
                start_row=start_row,
                start_column=start_col,
                end_row=start_row,
                end_column=start_col + ncols - 1,
            )

        for j, col_name in enumerate(cols):
            cell = ws.cell(start_row + 1, start_col + j, str(col_name))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for i, period in enumerate(periods):
            excel_row = start_row + 2 + i

            for j, col_name in enumerate(cols):
                cell = ws.cell(excel_row, start_col + j)
                cell.number_format = default_num_fmt
                cell.alignment = center
                cell.border = thin_border

                if str(period).upper() == "TOTAL":
                    cell.value = None
                    cell.font = total_font
                    cell.fill = white_fill
                    continue

                if col_name in formula_map:
                    cell.value = formula_map[col_name](excel_row)
                else:
                    val = df.loc[period, col_name]
                    cell.value = float(val) if not pd.isna(val) else 0.0

                cell.font = normal_font
                cell.fill = data_fill

        if "TOTAL" in periods:
            total_idx = periods.index("TOTAL")
            bottom_row = start_row + 2 + total_idx - 1
        else:
            bottom_row = start_row + 2 + nrows - 1

        if outline:
            outline_border(ws, start_row, start_col, bottom_row, start_col + ncols - 1)

        inject_total_formulas(ws, df, start_row, start_col)
        style_total_row(ws, df, start_row, start_col)

        return max(bottom_row, start_row + 2 + nrows - 1), start_col + ncols - 1

    # -------------------------
    # PC Disc layout helpers
    # -------------------------

    def split_intra_eu_spreads(df: pd.DataFrame, eu: List[str]) -> List[Tuple[str, pd.DataFrame]]:
        blocks = []

        for i, A in enumerate(eu):
            cols = [f"{A}/ {B}" for B in eu[i + 1:]]
            cols = [c for c in cols if c in df.columns]

            if cols:
                blocks.append((A, df[cols]))

        return blocks

    def split_cross_spreads(df: pd.DataFrame, eu: List[str], asia: List[str]) -> List[Tuple[str, pd.DataFrame]]:
        blocks = []

        for A in eu:
            cols = [f"{A}/ {B}" for B in asia]
            cols = [c for c in cols if c in df.columns]

            if cols:
                blocks.append((A, df[cols]))

        return blocks

    def create_west_east_placeholder(index: pd.Index) -> pd.DataFrame:
        out = pd.DataFrame(index=index)
        out["WEST"] = None
        out["EAST"] = None
        return out

    def create_total_placeholder(index: pd.Index) -> pd.DataFrame:
        out = pd.DataFrame(index=index)
        out["TOTAL"] = None
        return out

    def west_east_formula_map(
            west_first_col: Optional[int],
            west_last_col: Optional[int],
            east_first_col: Optional[int],
            east_last_col: Optional[int],
    ) -> Dict[str, callable]:

        def west_formula(row):
            if west_first_col is None or west_last_col is None:
                return "=0"

            return (
                f"=SUM({get_column_letter(west_first_col)}{row}:"
                f"{get_column_letter(west_last_col)}{row})"
            )

        def east_formula(row):
            if east_first_col is None or east_last_col is None:
                return "=0"

            return (
                f"=SUM({get_column_letter(east_first_col)}{row}:"
                f"{get_column_letter(east_last_col)}{row})"
            )

        return {
            "WEST": west_formula,
            "EAST": east_formula,
        }

    def total_formula_map(west_col_num: int, east_col_num: int) -> Dict[str, callable]:
        west_col = get_column_letter(west_col_num)
        east_col = get_column_letter(east_col_num)

        def total_formula(row):
            return f"={west_col}{row}+{east_col}{row}"

        return {
            "TOTAL": total_formula,
        }

    # -------------------------
    # Sheet writers
    # -------------------------

    def add_sheet_input_metadata(wb):
        ws = wb.create_sheet("Input Metadata")
        ws.sheet_view.showGridLines = False

        ws.cell(1, 1, "Input Metadata").font = title_font

        ws.cell(3, 1, "IFRS label")
        ws.cell(3, 2, ifrs_label)

        ws.cell(4, 1, "Source sheet")
        ws.cell(4, 2, source_sheet_name)

        ws.cell(5, 1, "Source year from D1")
        ws.cell(5, 2, source_year)

        ws.cell(7, 1, f"Product names from row {PRODUCT_HEADER_ROW}")
        ws.cell(7, 1).font = section_font

        ws.cell(7, 2, "Used as")
        ws.cell(7, 3, "Treatment")

        for i, product in enumerate(source_products or [], start=8):
            mapped = (rename_used or {}).get(product, product)

            ws.cell(i, 1, product)
            ws.cell(i, 2, mapped)

            if product in EXCLUDE_FROM_OUTPUT_PRODUCTS:
                treatment = "Excluded from output"
            elif is_freight_product(product):
                treatment = "Shipping / passthrough"
            elif mapped != product:
                treatment = "Mapped for netting"
            elif product in SCALE_DIVIDE_BY_MILLION:
                treatment = "Divided by 1,000,000"
            else:
                treatment = "Standard product"

            ws.cell(i, 3, treatment)

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 35

        ws.sheet_state = "hidden"

    def add_sheet_netting_only(wb, sheet_name: str, stage_raw: Dict[str, pd.DataFrame]):
        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "B5"

        stage = normalize_for_export(stage_raw)
        periods = list(stage["_master_index"].index.astype(str))

        eu = cfg.hubs_eu
        asia = cfg.hubs_asia

        extras = [c for c in stage["outrights"].columns if c not in eu + asia]

        for k in ["outrights", "pos_after_eu", "pos_after_asia", "residual"]:
            stage[k] = stage[k].drop(columns=extras, errors="ignore")

        ws.cell(1, 1, f"{sheet_name}   [{unit}]").font = title_font
        ws.cell(1, 1).alignment = left
        ws.cell(1, 1).fill = white_fill

        # =====================================================
        # BLOCK 1: OUTRIGHTS + EU SPREADS + ASIA SPREADS
        # =====================================================
        r0 = 3
        write_period_column(ws, start_row=r0 + 1, start_col=1, periods=periods)

        out = stage["outrights"]
        out_eu_cols = [c for c in eu if c in out.columns]
        out_asia_cols = [c for c in asia if c in out.columns]

        c = 2

        out_eu_start = out_eu_end = None
        out_asia_start = out_asia_end = None

        if out_eu_cols:
            out_eu_start = c
            _, out_eu_end = write_box(
                ws,
                "Outrights (EU)",
                out[out_eu_cols],
                r0,
                c,
                num_fmt_net,
                yellow_fill,
            )
            c = out_eu_end + 2

        if out_asia_cols:
            out_asia_start = c
            _, out_asia_end = write_box(
                ws,
                "Outrights (Asia)",
                out[out_asia_cols],
                r0,
                c,
                num_fmt_net,
                yellow_fill,
            )
            c = out_asia_end + 2

        summary_start = c
        out_summary = create_west_east_placeholder(out.index)

        _, summary_end = write_box(
            ws,
            "Outrights",
            out_summary,
            r0,
            summary_start,
            num_fmt_net,
            white_fill,
            west_east_formula_map(
                out_eu_start,
                out_eu_end,
                out_asia_start,
                out_asia_end,
            ),
        )

        total_start = summary_end + 2
        out_total = create_total_placeholder(out.index)

        _, total_end = write_box(
            ws,
            "TOTAL",
            out_total,
            r0,
            total_start,
            num_fmt_net,
            white_fill,
            total_formula_map(summary_start, summary_start + 1),
        )

        c = total_end + 2

        for hub, df_block in split_intra_eu_spreads(stage["eu_spreads"], eu):
            _, c2 = write_box(
                ws,
                f"{hub} Spreads",
                df_block,
                r0,
                c,
                num_fmt_net,
                white_fill,
            )
            c = c2 + 2

        if not stage["asia_spreads"].empty:
            _, c2 = write_box(
                ws,
                "Asia Spreads",
                stage["asia_spreads"],
                r0,
                c,
                num_fmt_net,
                white_fill,
            )
            c = c2 + 2

        # =====================================================
        # BLOCK 2: LEFTOVER + EU → ASIA SPREADS
        # =====================================================
        r1 = r0 + 2 + len(periods) + 4
        write_period_column(ws, start_row=r1 + 1, start_col=1, periods=periods)

        leftover = stage["pos_after_asia"]
        left_eu_cols = [c for c in eu if c in leftover.columns]
        left_asia_cols = [c for c in asia if c in leftover.columns]

        c = 2

        left_eu_start = left_eu_end = None
        left_asia_start = left_asia_end = None

        if left_eu_cols:
            left_eu_start = c
            _, left_eu_end = write_box(
                ws,
                "Leftover After EU/Asia Netting (EU)",
                leftover[left_eu_cols],
                r1,
                c,
                num_fmt_net,
                yellow_fill,
            )
            c = left_eu_end + 2

        if left_asia_cols:
            left_asia_start = c
            _, left_asia_end = write_box(
                ws,
                "Leftover After EU/Asia Netting (Asia)",
                leftover[left_asia_cols],
                r1,
                c,
                num_fmt_net,
                yellow_fill,
            )
            c = left_asia_end + 2

        summary_start = c
        left_summary = create_west_east_placeholder(leftover.index)

        _, summary_end = write_box(
            ws,
            "Leftover",
            left_summary,
            r1,
            summary_start,
            num_fmt_net,
            white_fill,
            west_east_formula_map(
                left_eu_start,
                left_eu_end,
                left_asia_start,
                left_asia_end,
            ),
        )

        total_start = summary_end + 2
        left_total = create_total_placeholder(leftover.index)

        _, total_end = write_box(
            ws,
            "TOTAL",
            left_total,
            r1,
            total_start,
            num_fmt_net,
            white_fill,
            total_formula_map(summary_start, summary_start + 1),
        )

        c = total_end + 2

        for hub, df_block in split_cross_spreads(stage["cross_spreads"], eu, asia):
            _, c2 = write_box(
                ws,
                f"{hub} → Asia",
                df_block,
                r1,
                c,
                num_fmt_net,
                white_fill,
            )
            c = c2 + 2

        # =====================================================
        # BLOCK 3: RESIDUAL LEFTOVER
        # =====================================================
        r2 = r1 + 2 + len(periods) + 4
        write_period_column(ws, start_row=r2 + 1, start_col=1, periods=periods)

        residual = stage["residual"]
        res_eu_cols = [c for c in eu if c in residual.columns]
        res_asia_cols = [c for c in asia if c in residual.columns]

        c = 2

        res_eu_start = res_eu_end = None
        res_asia_start = res_asia_end = None

        if res_eu_cols:
            res_eu_start = c
            _, res_eu_end = write_box(
                ws,
                "Residual Leftover (EU)",
                residual[res_eu_cols],
                r2,
                c,
                num_fmt_net,
                yellow_fill,
            )
            c = res_eu_end + 2

        if res_asia_cols:
            res_asia_start = c
            _, res_asia_end = write_box(
                ws,
                "Residual Leftover (Asia)",
                residual[res_asia_cols],
                r2,
                c,
                num_fmt_net,
                yellow_fill,
            )
            c = res_asia_end + 2

        summary_start = c
        residual_summary = create_west_east_placeholder(residual.index)

        _, summary_end = write_box(
            ws,
            "Residual",
            residual_summary,
            r2,
            summary_start,
            num_fmt_net,
            white_fill,
            west_east_formula_map(
                res_eu_start,
                res_eu_end,
                res_asia_start,
                res_asia_end,
            ),
        )

        total_start = summary_end + 2
        residual_total = create_total_placeholder(residual.index)

        write_box(
            ws,
            "TOTAL",
            residual_total,
            r2,
            total_start,
            num_fmt_net,
            white_fill,
            total_formula_map(summary_start, summary_start + 1),
        )

        set_col_widths(ws, max_col=160, default=8.5)

    def add_sheet_other_products(wb, sheet_name: str, stage_raw: Dict[str, pd.DataFrame]):
        stage = normalize_for_export(stage_raw)

        eu = cfg.hubs_eu
        asia = cfg.hubs_asia

        extras = [c for c in stage["outrights"].columns if c not in eu + asia]

        extras = [
            c for c in extras
            if c not in EXCLUDE_FROM_OUTPUT_PRODUCTS
               and not is_freight_product(c)
        ]

        if not extras:
            return

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "B5"

        periods = list(stage["_master_index"].index.astype(str))

        ws.cell(1, 1, f"{sheet_name}   [As Input]").font = title_font
        ws.cell(1, 1).alignment = left

        ws.cell(
            2,
            1,
            "Other products are exported as inputs. No unit conversion applied.",
        ).font = Font(size=8, italic=True)

        title_row = 3
        header_row = 4
        first_data_row = 5

        other_all = stage["outrights"][extras].copy()

        # Raw block
        raw_cols = [c for c in RAW_OTHER_PRODUCT_ORDER if c in other_all.columns]

        for c in other_all.columns:
            if (
                    c not in raw_cols
                    and c not in EXCLUDE_FROM_OUTPUT_PRODUCTS
                    and not is_freight_product(c)
            ):
                raw_cols.append(c)

        raw_other = other_all[raw_cols] if raw_cols else pd.DataFrame(index=other_all.index)

        raw_period_col = 1
        raw_start_col = 2
        raw_end_col = raw_start_col + len(raw_cols) - 1

        if raw_cols:
            ws.merge_cells(
                start_row=title_row,
                start_column=raw_start_col,
                end_row=title_row,
                end_column=raw_end_col,
            )
            title_cell = ws.cell(title_row, raw_start_col, "Other Products")
            title_cell.font = section_font
            title_cell.alignment = center
            title_cell.fill = white_fill
            title_cell.border = thin_border

        cell = ws.cell(header_row, raw_period_col, "Period")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

        for j, col_name in enumerate(raw_cols):
            cnum = raw_start_col + j
            cell = ws.cell(header_row, cnum, col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center
            cell.border = thin_border

        for i, period in enumerate(periods):
            r = first_data_row + i

            pcell = ws.cell(r, raw_period_col, period)
            pcell.font = total_font if str(period).upper() == "TOTAL" else period_font
            pcell.fill = white_fill
            pcell.alignment = left
            pcell.border = thin_border

            for j, col_name in enumerate(raw_cols):
                cnum = raw_start_col + j
                cell = ws.cell(r, cnum)
                cell.number_format = num_fmt_passthrough
                cell.alignment = center
                cell.border = thin_border
                cell.fill = raw_fill
                cell.font = total_font if str(period).upper() == "TOTAL" else normal_font

                if str(period).upper() == "TOTAL":
                    col_letter = get_column_letter(cnum)
                    cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{r - 1})"
                else:
                    val = raw_other.loc[period, col_name]
                    cell.value = float(val) if not pd.isna(val) else 0.0

        if raw_cols:
            outline_border(
                ws,
                title_row,
                raw_start_col,
                first_data_row + len(periods) - 2,
                raw_end_col,
            )

        # Oil view
        oil_period_col = raw_end_col + 3
        oil_start_col = oil_period_col + 1
        oil_main_cols = ["Brent", "JCC", "Dubai"]
        oil_total_col = oil_start_col + 3

        cell = ws.cell(header_row, oil_period_col, "")
        cell.fill = white_fill
        cell.border = thin_border

        for j, col_name in enumerate(oil_main_cols):
            cnum = oil_start_col + j
            cell = ws.cell(header_row, cnum, col_name)
            cell.font = header_font
            cell.fill = orange_fill
            cell.alignment = center
            cell.border = thin_border

        cell = ws.cell(header_row, oil_total_col, "Total")
        cell.font = header_font
        cell.fill = white_fill
        cell.alignment = center
        cell.border = thin_border

        raw_cols_actual = list(raw_other.columns)

        def raw_cell(product_name: str, row_num: int) -> str:
            if product_name not in raw_cols_actual:
                return "0"
            col_num = raw_start_col + raw_cols_actual.index(product_name)
            return f"{get_column_letter(col_num)}{row_num}"

        def m2_raw_cell(product_name: str, row_num: int) -> str:
            lag_row = row_num - 2
            if lag_row < first_data_row:
                return "0"
            return raw_cell(product_name, lag_row)

        for i, period in enumerate(periods):
            r = first_data_row + i

            pcell = ws.cell(r, oil_period_col, period)
            pcell.font = total_font if str(period).upper() == "TOTAL" else period_font
            pcell.fill = white_fill
            pcell.alignment = left
            pcell.border = thin_border

            if str(period).upper() == "TOTAL":
                for j in range(3):
                    cnum = oil_start_col + j
                    col_letter = get_column_letter(cnum)
                    cell = ws.cell(r, cnum)
                    cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{r - 1})"
                    cell.number_format = num_fmt_passthrough
                    cell.font = total_font
                    cell.fill = orange_fill
                    cell.alignment = center
                    cell.border = thin_border

                total_letter = get_column_letter(oil_total_col)
                cell = ws.cell(r, oil_total_col)
                cell.value = f"=SUM({total_letter}{first_data_row}:{total_letter}{r - 1})"
                cell.number_format = num_fmt_passthrough
                cell.font = total_font
                cell.fill = white_fill
                cell.alignment = center
                cell.border = thin_border

            else:
                brent_formula = (
                        "="
                        + raw_cell("Brent Bullet", r)
                        + "+"
                        + raw_cell("Brent Futures", r)
                        + "+"
                        + m2_raw_cell("Brent Swaps", r)
                        + "+"
                        + m2_raw_cell("Dated Brent", r)
                )

                jcc_formula = f"={raw_cell('JCC', r)}"
                dubai_formula = f"={raw_cell('Dubai', r)}"

                formulas = [brent_formula, jcc_formula, dubai_formula]

                for j, formula in enumerate(formulas):
                    cnum = oil_start_col + j
                    cell = ws.cell(r, cnum)
                    cell.value = formula
                    cell.number_format = num_fmt_passthrough
                    cell.font = normal_font
                    cell.fill = orange_fill
                    cell.alignment = center
                    cell.border = thin_border

                brent_col = get_column_letter(oil_start_col)
                dubai_col = get_column_letter(oil_start_col + 2)

                cell = ws.cell(r, oil_total_col)
                cell.value = f"=SUM({brent_col}{r}:{dubai_col}{r})"
                cell.number_format = num_fmt_passthrough
                cell.font = normal_font
                cell.fill = white_fill
                cell.alignment = center
                cell.border = thin_border

        outline_border(
            ws,
            header_row,
            oil_start_col,
            first_data_row + len(periods) - 2,
            oil_start_col + 2,
        )

        # JCC/B ratio block
        ratio_period_col = oil_total_col + 3
        ratio_start_col = ratio_period_col + 1

        ratio_cols = ["JCC/B ratio", "B eq.", "Brent+Dubai", "Net Eq."]

        for j, col_name in enumerate(ratio_cols):
            cnum = ratio_start_col + j
            cell = ws.cell(header_row, cnum, col_name)
            cell.font = header_font
            cell.alignment = center
            cell.border = thin_border

            if col_name == "JCC/B ratio":
                cell.fill = input_ratio_fill
            elif col_name == "Net Eq.":
                cell.fill = net_eq_fill
            else:
                cell.fill = white_fill

        for i, period in enumerate(periods):
            r = first_data_row + i

            pcell = ws.cell(r, ratio_period_col, period)
            pcell.font = total_font if str(period).upper() == "TOTAL" else period_font
            pcell.fill = white_fill
            pcell.alignment = left
            pcell.border = thin_border

            ratio_col_letter = get_column_letter(ratio_start_col)
            b_eq_col_letter = get_column_letter(ratio_start_col + 1)
            brent_dubai_col_letter = get_column_letter(ratio_start_col + 2)

            oil_brent_col = get_column_letter(oil_start_col)
            oil_jcc_col = get_column_letter(oil_start_col + 1)
            oil_dubai_col = get_column_letter(oil_start_col + 2)

            for j, col_name in enumerate(ratio_cols):
                cnum = ratio_start_col + j
                cell = ws.cell(r, cnum)
                cell.number_format = num_fmt_passthrough
                cell.alignment = center
                cell.border = thin_border
                cell.font = total_font if str(period).upper() == "TOTAL" else normal_font

                if col_name == "JCC/B ratio":
                    cell.fill = input_ratio_fill
                    cell.value = None

                elif col_name == "B eq.":
                    cell.fill = white_fill
                    if str(period).upper() == "TOTAL":
                        col_letter = get_column_letter(cnum)
                        cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{r - 1})"
                    else:
                        cell.value = f"={oil_jcc_col}{r}*{ratio_col_letter}{r}"

                elif col_name == "Brent+Dubai":
                    cell.fill = white_fill
                    if str(period).upper() == "TOTAL":
                        col_letter = get_column_letter(cnum)
                        cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{r - 1})"
                    else:
                        cell.value = f"={oil_brent_col}{r}+{oil_dubai_col}{r}"

                elif col_name == "Net Eq.":
                    cell.fill = net_eq_fill
                    if str(period).upper() == "TOTAL":
                        col_letter = get_column_letter(cnum)
                        cell.value = f"=SUM({col_letter}{first_data_row}:{col_letter}{r - 1})"
                    else:
                        cell.value = f"={b_eq_col_letter}{r}+{brent_dubai_col_letter}{r}"

        outline_border(
            ws,
            header_row,
            ratio_start_col,
            first_data_row + len(periods) - 2,
            ratio_start_col + len(ratio_cols) - 1,
        )

        set_col_widths(ws, max_col=120, default=11)

        ws.column_dimensions["A"].width = 12
        ws.column_dimensions[get_column_letter(oil_period_col)].width = 12
        ws.column_dimensions[get_column_letter(ratio_period_col)].width = 12

        for col_num in range(raw_start_col, raw_end_col + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 12

        for col_num in range(oil_start_col, oil_total_col + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 12

        for col_num in range(ratio_start_col, ratio_start_col + len(ratio_cols)):
            ws.column_dimensions[get_column_letter(col_num)].width = 12

    def add_sheet_shipping(wb, sheet_name: str, stage_raw: Dict[str, pd.DataFrame]):
        stage = normalize_for_export(stage_raw)

        eu = cfg.hubs_eu
        asia = cfg.hubs_asia

        extras = [c for c in stage["outrights"].columns if c not in eu + asia]
        freight_cols = [c for c in extras if is_freight_product(c)]

        if not freight_cols:
            return

        ws = wb.create_sheet(sheet_name)
        ws.sheet_view.showGridLines = False
        ws.sheet_view.zoomScale = 90
        ws.freeze_panes = "B5"

        periods = list(stage["_master_index"].index.astype(str))

        ws.cell(1, 1, f"{sheet_name}   [As Input]").font = title_font
        ws.cell(1, 1).alignment = left

        ws.cell(
            2,
            1,
            "Shipping are exported as inputs. No unit conversion applied.",
        ).font = Font(size=8, italic=True)

        r0 = 3

        write_period_column(
            ws,
            start_row=r0 + 1,
            start_col=1,
            periods=periods,
        )

        shipping_out = stage["outrights"][freight_cols]

        write_box(
            ws,
            "Shipping",
            shipping_out,
            start_row=r0,
            start_col=2,
            default_num_fmt=num_fmt_passthrough,
            data_fill=orange_fill,
        )

        set_col_widths(ws, max_col=30, default=14)

    # -------------------------
    # Workbook creation
    # -------------------------

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    add_sheet_netting_only(wb, f"PC Disc - M   {cob_date_str}", monthly)
    add_sheet_other_products(wb, f"Other Products - M   {cob_date_str}", monthly)
    add_sheet_shipping(wb, f"Shipping - M   {cob_date_str}", monthly)

    add_sheet_input_metadata(wb)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    return bio.getvalue()


# ============================================================
# 6) Streamlit dashboard
# ============================================================

st.set_page_config(page_title="Spread Netting (LNG Trading)", layout="wide")

st.title("Spread Netting (LNG Trading) — IFRS Discounted Inputs")

st.markdown(
    """
**Pipeline**
1) Upload the full LNG Exposure workbook.  
2) App reads these tabs: `IFRS 26 discounted`, `IFRS 27 discounted`, `IFRS 28 discounted`.  
3) One output file is created per IFRS tab.  
4) Monthly netting runs period-by-period for each IFRS file.  
5) `EUA` is excluded.  
6) `LNG H DES Med` maps to `ATL`.  
7) `PEG - DNK` maps to `PEG`; `ZTP - DNK` maps to `ZTP`.  
8) Shipping is exported separately.  
9) HH, Brent products, JCC and Dubai are divided by 1,000,000.  
"""
)

DEFAULT_EU = ["NWE", "ATL", "PVB", "NBP", "PEG", "ZTP", "PSV", "THE", "TTF"]
DEFAULT_ASIA = ["INDIA", "FE", "JKM"]

with st.sidebar:
    st.header("COB Date")

    cob_date = st.date_input(
        "COB date",
        value=default_cob_date,
        help="Used in output sheet names, file names, and stale-period highlighting.",
    )

    cob_date_str = pd.Timestamp(cob_date).strftime("%Y-%m-%d")

    st.divider()
    st.header("Netting Settings")

    st.caption("Hub order matters because netting is greedy.")

    eu_text = st.text_area("EU hubs", value=", ".join(DEFAULT_EU))
    asia_text = st.text_area("Asia hubs", value=", ".join(DEFAULT_ASIA))

    st.divider()
    st.header("Units")

    unit = st.radio(
        "Display / Export unit",
        options=["MMBtu", "TBtu"],
        index=1,
        help="Only netting hubs and spreads are converted. Other Products remain as prepared from IFRS tabs.",
    )

    st.divider()
    st.header("View")

    view_mode = st.radio(
        "Presentation mode",
        options=["Overview (Tables)", "Drilldown (Single period)"],
        index=0,
    )

uploaded = st.file_uploader("Upload full LNG Exposure Excel file (.xlsx)", type=["xlsx"])

if uploaded:
    eu_hubs = dedupe_preserve_order([
        _clean_header(x)
        for x in eu_text.split(",")
        if _clean_header(x)
    ])

    asia_hubs = dedupe_preserve_order([
        _clean_header(x)
        for x in asia_text.split(",")
        if _clean_header(x)
    ])

    eu_hubs = [h for h in eu_hubs if not _is_non_netted_product(h)]
    asia_hubs = [h for h in asia_hubs if not _is_non_netted_product(h)]

    cfg = NettingConfig(hubs_eu=eu_hubs, hubs_asia=asia_hubs)
    all_hubs = eu_hubs + asia_hubs

    processed_outputs = {}

    for ifrs_label, sheet_name in IFRS_SHEETS.items():
        try:
            df, source_year, source_products, rename_used, actual_sheet_name = read_ifrs_sheet_input(
                uploaded,
                sheet_name,
            )

            if df.empty:
                st.warning(f"{ifrs_label}: no usable rows found in `{sheet_name}`.")
                continue

            date_col = "Date"

            other_products = detect_other_products(df, date_col=date_col, nettable=all_hubs)

            for c in df.columns:
                if is_freight_product(c) and c not in other_products:
                    other_products.append(c)

            other_products = [
                p for p in other_products
                if p not in EXCLUDE_FROM_OUTPUT_PRODUCTS
            ]

            df = ensure_columns(df, all_hubs)

            all_inputs = all_hubs + other_products

            monthly_pos = monthly_positions_from_input(
                df,
                date_col=date_col,
                hubs=all_inputs,
                dayfirst=True,
            )

            if monthly_pos.empty:
                st.warning(f"{ifrs_label}: no non-zero monthly exposures after parsing.")
                continue

            m_stage = netting_outputs_for_periods_staged(monthly_pos, cfg)
            m_u = convert_stage_selective(m_stage, cfg, unit)

            processed_outputs[ifrs_label] = {
                "stage": m_u,
                "source_year": source_year,
                "source_products": source_products,
                "rename_used": rename_used,
                "actual_sheet_name": actual_sheet_name,
                "other_products": other_products,
                "monthly_pos": monthly_pos,
            }

        except Exception as e:
            st.error(f"{ifrs_label}: failed to process `{sheet_name}`.")
            st.exception(e)

    if not processed_outputs:
        st.error("No IFRS sheets were processed successfully.")
        st.stop()

    st.success(f"Processed {len(processed_outputs)} IFRS sheet(s): {list(processed_outputs.keys())}")

    preview_tabs = st.tabs(list(processed_outputs.keys()))


    def show_overview(stage_u: Dict[str, pd.DataFrame], label_prefix: str, passthrough_cols: List[str]):
        st.subheader(f"{label_prefix} — Outrights")
        st.dataframe(
            style_table_mixed(blank_zeros(stage_u["outrights"]), unit, passthrough_cols),
            use_container_width=True,
        )

        st.subheader(f"{label_prefix} — EU Spreads")
        st.dataframe(
            style_table_mixed(blank_zeros(stage_u["eu_spreads"]), unit, []),
            use_container_width=True,
        )

        st.subheader(f"{label_prefix} — Positions After EU Netting")
        st.dataframe(
            style_table_mixed(blank_zeros(stage_u["pos_after_eu"]), unit, passthrough_cols),
            use_container_width=True,
        )

        st.subheader(f"{label_prefix} — Asia Spreads")
        st.dataframe(
            style_table_mixed(blank_zeros(stage_u["asia_spreads"]), unit, []),
            use_container_width=True,
        )

        st.subheader(f"{label_prefix} — Positions After Asia Netting")
        st.dataframe(
            style_table_mixed(blank_zeros(stage_u["pos_after_asia"]), unit, passthrough_cols),
            use_container_width=True,
        )

        st.subheader(f"{label_prefix} — EU → Asia Spreads")
        st.dataframe(
            style_table_mixed(blank_zeros(stage_u["cross_spreads"]), unit, []),
            use_container_width=True,
        )

        st.subheader(f"{label_prefix} — Residual Outrights")
        st.dataframe(
            style_table_mixed(blank_zeros(stage_u["residual"]), unit, passthrough_cols),
            use_container_width=True,
        )


    def show_drilldown(stage_u: Dict[str, pd.DataFrame], period_label: str, passthrough_cols: List[str]):
        periods = list(stage_u["outrights"].index.astype(str))
        chosen = st.selectbox(period_label, options=periods)
        idx = stage_u["outrights"].index[periods.index(chosen)]

        col_a, col_b = st.columns([1, 1])

        with col_a:
            st.markdown("### Outrights")
            s = nonzero_only_row(stage_u["outrights"].loc[idx])
            st.dataframe(
                style_table_mixed(blank_zeros(s.to_frame("Value").T), unit, passthrough_cols),
                use_container_width=True,
            )

            st.markdown("### Positions After EU Netting")
            s = nonzero_only_row(stage_u["pos_after_eu"].loc[idx])
            st.dataframe(
                style_table_mixed(blank_zeros(s.to_frame("Value").T), unit, passthrough_cols),
                use_container_width=True,
            )

            st.markdown("### Positions After Asia Netting")
            s = nonzero_only_row(stage_u["pos_after_asia"].loc[idx])
            st.dataframe(
                style_table_mixed(blank_zeros(s.to_frame("Value").T), unit, passthrough_cols),
                use_container_width=True,
            )

            st.markdown("### Residual")
            s = nonzero_only_row(stage_u["residual"].loc[idx])
            st.dataframe(
                style_table_mixed(blank_zeros(s.to_frame("Value").T), unit, passthrough_cols),
                use_container_width=True,
            )

        with col_b:
            st.markdown("### EU Spreads")
            s = nonzero_only_row(stage_u["eu_spreads"].loc[idx])
            st.dataframe(
                style_table_mixed(blank_zeros(s.to_frame("Value").T), unit, []),
                use_container_width=True,
            )

            st.markdown("### Asia Spreads")
            s = nonzero_only_row(stage_u["asia_spreads"].loc[idx])
            st.dataframe(
                style_table_mixed(blank_zeros(s.to_frame("Value").T), unit, []),
                use_container_width=True,
            )

            st.markdown("### EU → Asia Spreads")
            s = nonzero_only_row(stage_u["cross_spreads"].loc[idx])
            st.dataframe(
                style_table_mixed(blank_zeros(s.to_frame("Value").T), unit, []),
                use_container_width=True,
            )


    for tab, ifrs_label in zip(preview_tabs, processed_outputs.keys()):
        with tab:
            obj = processed_outputs[ifrs_label]
            st.info(
                f"{ifrs_label}: source sheet `{obj['actual_sheet_name']}`. "
                f"Source D1 year/date parsed as `{obj['source_year']}`."
            )

            if view_mode == "Overview (Tables)":
                show_overview(obj["stage"], ifrs_label, obj["other_products"])
            else:
                show_drilldown(obj["stage"], f"{ifrs_label} month", obj["other_products"])

    st.divider()
    st.subheader("Download IFRS Excel files")

    for ifrs_label, obj in processed_outputs.items():
        try:
            xbytes = export_to_excel_bytes_staged(
                monthly=obj["stage"],
                unit=unit,
                cfg=cfg,
                ifrs_label=ifrs_label,
                source_sheet_name=obj["actual_sheet_name"],
                cob_date_str=cob_date_str,
                source_year=obj["source_year"],
                source_products=obj["source_products"],
                rename_used=obj["rename_used"],
            )

            safe_label = ifrs_label.replace(" ", "_")
            file_name = f"{ifrs_label} COB {cob_date_str}.xlsx"

            st.download_button(
                label=f"Download {ifrs_label} COB {cob_date_str}.xlsx",
                data=xbytes,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"download_{safe_label}_{unit}_{cob_date_str}",
            )

        except Exception as e:
            st.error(f"Excel generation failed for {ifrs_label}: {e}")
            st.exception(e)

else:
    st.info("Upload the full LNG Exposure workbook to generate IFRS 26, IFRS 27 and IFRS 28 outputs.")
    st.code(
        "pip install pandas openpyxl streamlit\n"
        "streamlit run \"Trying to automate spread netting.py\"",
        language="bash",
    )

